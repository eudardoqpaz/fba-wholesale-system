"""
Excel Export Service - Generate formatted Excel files with colors and styles.
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Color definitions
HEADER_FILL = PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid")  # Amazon orange
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="000000")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FF9900")
SUBTITLE_FONT = Font(name="Calibri", size=10, color="666666")
MONEY_FORMAT = '$#,##0.00'
PERCENT_FORMAT = '0.0%'
NUMBER_FORMAT = '#,##0'

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
LIGHT_GRAY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

GREEN_FONT = Font(name="Calibri", size=10, color="006100")
YELLOW_FONT = Font(name="Calibri", size=10, color="9C6500")
RED_FONT = Font(name="Calibri", size=10, color="9C0006")
NORMAL_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", size=10, bold=True)
LINK_FONT = Font(name="Calibri", size=10, color="0563C1", underline="single")

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def generate_shopping_list_excel(shopping_list: list, summary: dict) -> io.BytesIO:
    """Generate a formatted Excel file with the shopping list."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Lista de Compras"

    # Column widths
    widths = {
        'A': 5,   # #
        'B': 14,  # ASIN
        'C': 45,  # Producto
        'D': 12,  # Marca
        'E': 12,  # Precio Amazon
        'F': 12,  # Costo Estimado
        'G': 10,  # Cantidad
        'H': 12,  # Inversion
        'I': 12,  # Ganancia
        'J': 8,   # ROI
        'K': 10,  # BSR
        'L': 10,  # Ventas/mes
        'M': 8,   # Riesgo
        'N': 30,  # Links
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Title
    ws.merge_cells('A1:N1')
    ws['A1'] = "LISTA DE COMPRAS - FBA WHOLESALE"
    ws['A1'].font = Font(name="Calibri", size=16, bold=True, color="FF9900")
    ws['A1'].alignment = Alignment(horizontal="center")

    # Summary
    row = 3
    summary_items = [
        ("Productos:", summary.get('total_products', 0)),
        ("Inversion Total:", f"${summary.get('total_investment', 0):.2f}"),
        ("Ganancia Esperada:", f"${summary.get('expected_profit', 0):.2f}"),
        ("ROI Esperado:", f"{summary.get('expected_roi', 0)}%"),
    ]
    for label, value in summary_items:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = BOLD_FONT
        ws[f'B{row}'] = value
        ws[f'B{row}'].font = Font(name="Calibri", size=11, bold=True, color="FF9900")
        row += 1

    # Headers
    row += 1
    headers = ["#", "ASIN", "Producto", "Marca", "Precio Amazon", "Costo Est.", "Cantidad",
               "Inversion", "Ganancia", "ROI %", "BSR", "Sales/mo", "Riesgo", "Links"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # Data rows
    for i, product in enumerate(shopping_list):
        row += 1
        calc = product.get('calc', {})
        risk = product.get('risk', 'MEDIO')

        # Alternating row colors
        fill = LIGHT_GRAY_FILL if i % 2 == 0 else WHITE_FILL

        # Row data
        values = [
            i + 1,
            product.get('asin', ''),
            product.get('title', '')[:60],
            product.get('brand', ''),
            product.get('amazon_price', 0),
            product.get('estimated_wholesale_cost', product.get('supplier_cost', 0)),
            product.get('recommended_qty', product.get('quantity', 0)),
            product.get('total_cost', 0),
            product.get('expected_profit', calc.get('net_profit', 0)),
            product.get('roi_pct', calc.get('roi_pct', 0)),
            product.get('bsr', 0),
            product.get('monthly_sales', product.get('monthly_sales_est', 0)),
            risk,
            f"https://www.amazon.com/dp/{product.get('asin', '')}",
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.fill = fill
            cell.border = THIN_BORDER

            # Format specific columns
            if col in [5, 6, 8, 9]:  # Money columns
                cell.number_format = MONEY_FORMAT
                cell.alignment = Alignment(horizontal="right")
            elif col == 10:  # ROI
                cell.number_format = '0.0"%"'
                cell.alignment = Alignment(horizontal="center")
            elif col == 11:  # BSR
                cell.number_format = NUMBER_FORMAT
                cell.alignment = Alignment(horizontal="right")
            elif col == 12:  # Sales
                cell.number_format = NUMBER_FORMAT
                cell.alignment = Alignment(horizontal="right")
            elif col == 7:  # Quantity
                cell.alignment = Alignment(horizontal="center")
                cell.font = BOLD_FONT

            # Color coding for ROI
            if col == 10:
                roi_val = value if isinstance(value, (int, float)) else 0
                if roi_val >= 25:
                    cell.fill = GREEN_FILL
                    cell.font = GREEN_FONT
                elif roi_val >= 15:
                    cell.fill = YELLOW_FILL
                    cell.font = YELLOW_FONT
                else:
                    cell.fill = RED_FILL
                    cell.font = RED_FONT

            # Color coding for Risk
            if col == 13:
                if value == "BAJO":
                    cell.fill = GREEN_FILL
                    cell.font = GREEN_FONT
                elif value == "MEDIO":
                    cell.fill = YELLOW_FILL
                    cell.font = YELLOW_FONT
                else:
                    cell.fill = RED_FILL
                    cell.font = RED_FONT

            # Make link clickable
            if col == 14 and value.startswith("http"):
                cell.font = LINK_FONT
                cell.hyperlink = value
                cell.value = "Ver en Amazon"

    # Totals row
    row += 1
    ws.cell(row=row, column=6, value="TOTALES:").font = BOLD_FONT
    ws.cell(row=row, column=6).alignment = Alignment(horizontal="right")

    total_investment = sum(p.get('total_cost', 0) for p in shopping_list)
    total_profit = sum(p.get('expected_profit', p.get('calc', {}).get('net_profit', 0)) for p in shopping_list)
    total_qty = sum(p.get('recommended_qty', p.get('quantity', 0)) for p in shopping_list)

    ws.cell(row=row, column=7, value=total_qty).font = BOLD_FONT
    ws.cell(row=row, column=7).alignment = Alignment(horizontal="center")

    ws.cell(row=row, column=8, value=total_investment).font = Font(name="Calibri", size=11, bold=True, color="FF9900")
    ws.cell(row=row, column=8).number_format = MONEY_FORMAT

    ws.cell(row=row, column=9, value=total_profit).font = Font(name="Calibri", size=11, bold=True, color="006100")
    ws.cell(row=row, column=9).number_format = MONEY_FORMAT

    # Store Links Sheet
    ws2 = wb.create_sheet("Links de Compra")
    ws2.column_dimensions['A'].width = 14
    ws2.column_dimensions['B'].width = 40
    ws2.column_dimensions['C'].width = 20
    ws2.column_dimensions['D'].width = 20
    ws2.column_dimensions['E'].width = 20
    ws2.column_dimensions['F'].width = 20
    ws2.column_dimensions['G'].width = 20
    ws2.column_dimensions['H'].width = 20

    ws2.merge_cells('A1:H1')
    ws2['A1'] = "LINKS DIRECTOS PARA COMPRAR"
    ws2['A1'].font = TITLE_FONT
    ws2['A1'].alignment = Alignment(horizontal="center")

    link_headers = ["ASIN", "Producto", "Amazon", "Walmart", "Costco", "Sam's Club", "Faire", "Google Shopping"]
    for col, header in enumerate(link_headers, 1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    for i, product in enumerate(shopping_list):
        row = i + 4
        asin = product.get('asin', '')
        title = product.get('title', '')[:50]
        search = title.replace(' ', '+')

        ws2.cell(row=row, column=1, value=asin).font = BOLD_FONT
        ws2.cell(row=row, column=2, value=title).font = NORMAL_FONT

        links = [
            f"https://www.amazon.com/dp/{asin}",
            f"https://www.walmart.com/search?q={search}",
            f"https://www.costco.com/CatalogSearch?keyword={search}",
            f"https://www.samsclub.com/search/{search}",
            f"https://www.faire.com/search?q={search}",
            f"https://www.google.com/search?q={search}+wholesale+buy&tbm=shop",
        ]

        for col, link in enumerate(links, 3):
            cell = ws2.cell(row=row, column=col, value="Click para ir")
            cell.font = LINK_FONT
            cell.hyperlink = link
            cell.border = THIN_BORDER
            if row % 2 == 0:
                cell.fill = LIGHT_GRAY_FILL

    # Freeze panes
    ws.freeze_panes = 'A7'
    ws2.freeze_panes = 'A4'

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_scan_results_excel(products: list, stats: dict) -> io.BytesIO:
    """Generate Excel from scan/finder results."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"

    # Column widths
    widths = {'A': 14, 'B': 45, 'C': 15, 'D': 12, 'E': 10, 'F': 8, 'G': 12, 'H': 10, 'I': 8, 'J': 8, 'K': 25}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Title
    ws.merge_cells('A1:K1')
    ws['A1'] = "RESULTADOS DEL ESCANEO - FBA WHOLESALE"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal="center")

    # Stats
    row = 3
    stats_items = [
        ("Total Escaneados:", stats.get('total_scanned', 0)),
        ("Rentables:", stats.get('profitable_count', stats.get('profitable_found', 0))),
        ("Marginales:", stats.get('marginal_count', stats.get('marginal_found', 0))),
        ("No Rentables:", stats.get('not_profitable_count', 0)),
        ("ROI Promedio:", f"{stats.get('avg_roi', 0)}%"),
    ]
    for label, value in stats_items:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = BOLD_FONT
        ws[f'B{row}'] = value
        ws[f'B{row}'].font = Font(name="Calibri", size=11, bold=True, color="FF9900")
        row += 1

    # Headers
    row += 1
    headers = ["ASIN", "Producto", "Precio Amazon", "Costo", "ROI %", "Ganancia", "BSR", "Sellers", "Sales/mo", "Riesgo", "Amazon Link"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # Data
    for i, p in enumerate(products):
        row += 1
        calc = p.get('calc', {})
        fill = LIGHT_GRAY_FILL if i % 2 == 0 else WHITE_FILL

        values = [
            p.get('asin', ''),
            p.get('title', '')[:60],
            p.get('amazon_price', p.get('sell_price', 0)),
            p.get('supplier_cost', p.get('estimated_wholesale_cost', 0)),
            calc.get('roi_pct', p.get('roi_pct', 0)),
            calc.get('net_profit', p.get('net_profit', 0)),
            p.get('bsr', 0),
            p.get('fba_seller_count', p.get('sellers', 0)),
            p.get('monthly_sales_est', p.get('monthly_sales', 0)),
            p.get('risk', ''),
            f"https://www.amazon.com/dp/{p.get('asin', '')}",
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.fill = fill
            cell.border = THIN_BORDER

            if col in [3, 4, 6]:
                cell.number_format = MONEY_FORMAT
            elif col == 5:
                cell.number_format = '0.0"%"'
                if isinstance(value, (int, float)):
                    if value >= 25:
                        cell.fill = GREEN_FILL
                        cell.font = GREEN_FONT
                    elif value >= 15:
                        cell.fill = YELLOW_FILL
                        cell.font = YELLOW_FONT
                    else:
                        cell.fill = RED_FILL
                        cell.font = RED_FONT
            elif col == 7:
                cell.number_format = NUMBER_FORMAT
            elif col == 11:
                cell.font = LINK_FONT
                cell.hyperlink = value
                cell.value = "Ver en Amazon"

    ws.freeze_panes = 'A6'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
